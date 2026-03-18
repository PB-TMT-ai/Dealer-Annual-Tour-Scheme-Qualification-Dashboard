"""
Generate Dealer Slab Upgrade Analysis Excel Report.

Produces a 3-tab Excel file categorising dealers by how much additional
lifting (MT) they need in the current month to move into the next slab.

Output: .workspace/Dealer_Slab_Upgrade_Analysis.xlsx
"""

import datetime as dt
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data"
OUTPUT_DIR = PROJECT_ROOT / ".workspace"
OUTPUT_FILE = OUTPUT_DIR / "Dealer_Slab_Upgrade_Analysis.xlsx"

SHEET_NAME = "FY 26_qualification scenario"
HEADER_ROW = 2

# ---------------------------------------------------------------------------
# Slab config (mirrored from app.py — single source of truth)
# ---------------------------------------------------------------------------
SLAB_CONFIG = [
    {"slab": "A", "lower": 200.0, "upper": 300.0, "gift": "Domestic Trip 1 pax 3N-4D"},
    {"slab": "B", "lower": 300.0, "upper": 500.0, "gift": "Intl Trip (1 pax South Asia) 3N-4D"},
    {"slab": "C", "lower": 500.0, "upper": 750.0, "gift": "Intl Trip Almaty / 2 pax SE Asia 3N-4D"},
    {"slab": "D", "lower": 750.0, "upper": 1250.0, "gift": "Intl Trip 2 pax SE Asia + EV / 2 pax Almaty"},
    {"slab": "E", "lower": 1250.0, "upper": float("inf"), "gift": "Luxury Intl Trip (2 pax Dubai) 3N-4D"},
]

NEXT_SLAB_MAP = {"No Slab": "A", "A": "B", "B": "C", "C": "D", "D": "E", "E": "Max Slab"}
NEXT_SLAB_THRESHOLD = {"No Slab": 200.0, "A": 300.0, "B": 500.0, "C": 750.0, "D": 1250.0}
SLAB_GIFT_MAP = {s["slab"]: s["gift"] for s in SLAB_CONFIG}
SLAB_GIFT_MAP["No Slab"] = "Non-Qualifier"

# Month column rename mapping (Excel datetime headers → short labels)
_MONTH_RENAME = {
    dt.datetime(2025, 4, 1): "Apr",
    dt.datetime(2025, 5, 1): "May",
    dt.datetime(2025, 6, 1): "Jun",
    dt.datetime(2025, 7, 1): "Jul",
    dt.datetime(2025, 8, 1): "Aug",
    dt.datetime(2025, 9, 1): "Sep",
    dt.datetime(2025, 10, 1): "Oct",
    dt.datetime(2025, 11, 1): "Nov",
    "Dec-25": "Dec",
    dt.datetime(2026, 1, 1): "Jan",
    dt.datetime(2026, 2, 1): "Feb",
    dt.datetime(2026, 3, 1): "Mar",
}

MONTH_LABELS = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]


# ---------------------------------------------------------------------------
# Helpers (from app.py)
# ---------------------------------------------------------------------------
def assign_slab(vol: float) -> str:
    if pd.isna(vol) or vol < 200:
        return "No Slab"
    if vol < 300:
        return "A"
    if vol < 500:
        return "B"
    if vol < 750:
        return "C"
    if vol < 1250:
        return "D"
    return "E"


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def load_data() -> pd.DataFrame:
    xlsx_files = sorted(DATA_DIR.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    if not xlsx_files:
        raise FileNotFoundError("No .xlsx file found in data/ folder")
    data_path = xlsx_files[0]
    print(f"Loading: {data_path.name}")

    raw = pd.read_excel(data_path, sheet_name=SHEET_NAME, header=HEADER_ROW)

    # Rename month columns
    rename_map: dict = {}
    for target, label in _MONTH_RENAME.items():
        for col in raw.columns:
            match = False
            if isinstance(target, dt.datetime) and isinstance(col, (dt.datetime, pd.Timestamp)):
                match = col == target
            elif isinstance(target, str) and isinstance(col, str):
                match = col == target
            if match:
                rename_map[col] = label
                break
    raw.rename(columns=rename_map, inplace=True)

    # Numeric coercion
    for col in MONTH_LABELS:
        if col in raw.columns:
            raw[col] = pd.to_numeric(raw[col], errors="coerce").fillna(0.0)
    raw["FY 26 total"] = pd.to_numeric(raw["FY 26 total"], errors="coerce").fillna(0.0)
    raw["Avg. monthly vol."] = pd.to_numeric(raw.get("Avg. monthly vol."), errors="coerce").fillna(0.0)

    # Clean text columns
    for col in ["State", "Zone"]:
        raw[col] = raw[col].apply(
            lambda x: "Unknown"
            if (isinstance(x, (int, float)) and not pd.isna(x)) or pd.isna(x)
            else str(x).strip().title() if col == "State" else str(x).strip()
        )
        raw.loc[raw[col].isin(["0", "0.0"]), col] = "Unknown"

    raw["District"] = raw["District"].apply(
        lambda x: "" if pd.isna(x) or (isinstance(x, (int, float)) and x == 0) else str(x).strip()
    )
    raw["Distributor Name"] = raw["Distributor Name"].apply(
        lambda x: "" if pd.isna(x) else str(x).strip()
    )

    # Derived columns
    raw["Current Slab"] = raw["FY 26 total"].apply(assign_slab)
    raw["Lifting Frequency"] = raw[MONTH_LABELS].gt(0).sum(axis=1).astype(int)
    raw["Upgrade Slab"] = raw["Current Slab"].map(NEXT_SLAB_MAP)
    raw["Lifting Required (MT)"] = raw.apply(
        lambda r: max(NEXT_SLAB_THRESHOLD.get(r["Current Slab"], 0) - r["FY 26 total"], 0)
        if r["Current Slab"] != "E" else None,
        axis=1,
    )
    raw["Current Gift"] = raw["Current Slab"].map(SLAB_GIFT_MAP)
    raw["Upgrade Gift"] = raw["Upgrade Slab"].map(SLAB_GIFT_MAP).fillna("")

    return raw


# ---------------------------------------------------------------------------
# Filter & split
# ---------------------------------------------------------------------------
OUTPUT_COLS = [
    "Sr No.",
    "Name of the Dealer",
    "Distributor Name",
    "State",
    "District",
    "Zone",
    "Dealer Segmentation",
    "Account Owner As per SF",
    "TM",
    "FY 26 total",
    "Mar",
    "Avg. monthly vol.",
    "Lifting Frequency",
    "Current Slab",
    "Upgrade Slab",
    "Lifting Required (MT)",
    "Current Gift",
    "Upgrade Gift",
]

DISPLAY_NAMES = {
    "Name of the Dealer": "Dealer Name",
    "Account Owner As per SF": "Account Owner",
    "FY 26 total": "FY 26 Total (MT)",
    "Mar": "March Lifting (MT)",
    "Avg. monthly vol.": "Avg Monthly Vol (MT)",
}


def build_report(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    # Exclude Slab E (already max) and dealers with no gap
    eligible = df[
        (df["Current Slab"] != "E")
        & (df["Lifting Required (MT)"].notna())
        & (df["Lifting Required (MT)"] > 0)
    ].copy()

    # No Slab filter: only include those needing ≤100 MT (FY 26 total ≥ 100)
    no_slab_mask = eligible["Current Slab"] == "No Slab"
    eligible = eligible[~no_slab_mask | (eligible["Lifting Required (MT)"] <= 100)]

    # Select and rename columns
    report = eligible[OUTPUT_COLS].copy()
    report.rename(columns=DISPLAY_NAMES, inplace=True)

    # Round numeric columns
    for col in ["FY 26 Total (MT)", "March Lifting (MT)", "Avg Monthly Vol (MT)", "Lifting Required (MT)"]:
        if col in report.columns:
            report[col] = report[col].round(1)

    # Split into 3 tabs
    tabs = {
        "50+ MT Required": report[report["Lifting Required (MT)"] >= 50]
            .sort_values("Lifting Required (MT)", ascending=True)
            .reset_index(drop=True),
        "20-50 MT Required": report[
            (report["Lifting Required (MT)"] >= 20) & (report["Lifting Required (MT)"] < 50)
        ]
            .sort_values("Lifting Required (MT)", ascending=True)
            .reset_index(drop=True),
        "0-20 MT Required": report[report["Lifting Required (MT)"] < 20]
            .sort_values("Lifting Required (MT)", ascending=True)
            .reset_index(drop=True),
    }

    return tabs


# ---------------------------------------------------------------------------
# Excel formatting
# ---------------------------------------------------------------------------
TAB_COLORS = {"50+ MT Required": "FF6B6B", "20-50 MT Required": "FFB347", "0-20 MT Required": "77DD77"}

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_worksheet(ws: "openpyxl.worksheet.worksheet.Worksheet") -> None:
    """Apply formatting to a worksheet."""
    # Style header row
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Auto-width columns
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 35)

    # Freeze header row
    ws.freeze_panes = "A2"


def write_excel(tabs: dict[str, pd.DataFrame]) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for sheet_name, df in tabs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Format each sheet
        for sheet_name in tabs:
            ws = writer.sheets[sheet_name]
            style_worksheet(ws)
            if sheet_name in TAB_COLORS:
                ws.sheet_properties.tabColor = TAB_COLORS[sheet_name]

    return OUTPUT_FILE


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    print("=" * 60)
    print("Dealer Slab Upgrade Analysis Report")
    print("=" * 60)

    df = load_data()
    print(f"Total dealers loaded: {len(df)}")
    print(f"Slab distribution: {df['Current Slab'].value_counts().to_dict()}")

    tabs = build_report(df)
    for name, tab_df in tabs.items():
        print(f"  {name}: {len(tab_df)} dealers")

    output_path = write_excel(tabs)
    print(f"\nReport saved to: {output_path}")
    print("Done!")


if __name__ == "__main__":
    main()
